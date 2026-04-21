import json
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment

with open('user_stats.json', 'r', encoding='utf-8') as f:
    boxlive_dict = json.load(f)

user_list = []
total_box = 0
total_cost = 0
total_profit = 0

for uid, user_info in boxlive_dict.items():
    count = user_info["count"]
    cost = user_info["cost"]
    profit = user_info["profit"]
    
    user_list.append({
        "用户名": user_info["uname"],
        "盲盒数": count,
        "总花费（电池）": round(cost*10),
        "总收益（电池）": round(profit*10),
        "净收益（电池）": round(profit*10)-round(cost*10)
    })
    
    total_box += count
    total_cost += cost
    total_profit += profit

total_cost *= 10
total_profit *= 10

df = pd.DataFrame(user_list)
date = input("请输入日期(yymmdd): ")
folder = "analysis"
os.makedirs(folder, exist_ok=True)
save_path = os.path.join(folder, f"{date}盲盒统计.xlsx")

df.to_excel(save_path, index=False)

wb = load_workbook(save_path)
ws = wb.active

ws.insert_rows(1)
ws.merge_cells('A1:E1')
ws['A1'] = f'总盲盒数：{total_box}，共花费{round(total_cost)}电池，共开出{round(total_profit)}电池，净收益{round(total_profit)-round(total_cost)}电池'
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15

wb.save(save_path)