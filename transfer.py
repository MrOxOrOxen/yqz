import json
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from datetime import datetime
import matplotlib.dates as mdates
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import sys
import jieba

if len(sys.argv) > 1:
    date = sys.argv[1]
else:
    date = datetime.now().strftime('%Y-%m-%d')

plt.rcParams['font.sans-serif'] = ['WenQuanYi Zen Hei', 'Microsoft YaHei', 'SimHei']
plt.rcParams['axes.unicode_minus'] = False

# date = input("请输入日期 (yyyy-mm-dd): ")
folder_box = "analysis/box"
folder_gift = "analysis/gift"
folder_all = "analysis/all"
os.makedirs(folder_box, exist_ok=True)
os.makedirs(folder_gift, exist_ok=True)
os.makedirs(folder_all, exist_ok=True)

# judgment
try:
    with open('files/danmu.jsonl', 'r', encoding='utf-8') as f:
        lines = f.readlines()
        if len(lines) <= 1:
            print("No live info. Program exiting.")
            sys.exit(1)
except Exception as e:
    print("No live info. Program exiting.")
    sys.exit(1)

# all.json
try:
    with open('files/all.json', 'r', encoding='utf-8') as f:
        all_data = json.load(f)

    stats = {}
    for item in all_data:
        price = item.get('gift_price', 0) 

        period = (item['time'] // 60) * 60 
        stats[period] = stats.get(period, 0) + price

    if stats:
        sorted_keys = sorted(stats.keys())
        min_time = sorted_keys[0]
        max_time = sorted_keys[-1]
        full_times = range(min_time, max_time + 60, 60)
        times = [datetime.fromtimestamp(t) for t in full_times]
        counts = [stats.get(t, 0) for t in full_times]

        plt.figure(figsize=(12, 6))
        
        plt.plot(times, counts, color='#ff85c0', linewidth=2.5, alpha=0.9, label='电池流水')
        plt.fill_between(times, counts, color='#ff85c0', alpha=0.3)

        ax = plt.gca()
        ax.xaxis.set_major_locator(mdates.MinuteLocator(byminute=range(0, 60, 20)))
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))

        plt.title(f'{date} 云宝直播间电池统计图')
        plt.ylabel('电池数量')
        plt.grid(axis='y', linestyle=':', alpha=0.6)
        plt.tight_layout()

        plt.savefig(f'{folder_all}/{date}电池统计.png', dpi=300)
    else:
        print(f"Warning: {date} all.json data is empty.")

except FileNotFoundError:
    print("Error: no all.json")

# gift.json
try:
    with open('files/gift.json', 'r', encoding='utf-8') as f:
        gift_data = json.load(f)

    gift_list = []
    all_gift_profit = 0
    
    for uid, info in gift_data.items():
        details = ", ".join([f"{name}*{num}" for name, num in info['gift_list'].items()])
        profit = info['profit']
        all_gift_profit += profit
        
        gift_list.append({
            "用户名": info['uname'],
            "电池总数": profit,
            "礼物详情": details
        })

    gift_df = pd.DataFrame(gift_list)
    gift_df = gift_df.sort_values(by="电池总数", ascending=False)

    gift_path = os.path.join(folder_gift, f"{date}礼物详情统计.xlsx")
    gift_df.to_excel(gift_path, index=False)

    wb_gift = load_workbook(gift_path)
    ws_gift = wb_gift.active
    
    ws_gift.insert_rows(1)
    ws_gift.merge_cells('A1:C1')
    ws_gift['A1'] = f'总收益：{round(all_gift_profit, 1)}电池'
    ws_gift['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws_gift.column_dimensions['A'].width = 25
    ws_gift.column_dimensions['B'].width = 15
    ws_gift.column_dimensions['C'].width = 70
    wb_gift.save(gift_path)
except FileNotFoundError:
    print("Error: no gift.json")

# box.json
try:
    with open('files/box.json', 'r', encoding='utf-8') as f:
        boxlive_dict = json.load(f)

    user_list = []
    total_box = 0
    total_cost = 0
    total_profit = 0

    for uid, user_info in boxlive_dict.items():
        count = user_info["count"]
        cost = user_info["cost"]
        profit = user_info["profit"]
        
        user_list.append({
            "用户名": user_info["uname"],
            "盲盒数": count,
            "总花费（电池）": round(cost),
            "总收益（电池）": round(profit),
            "净收益（电池）": round(profit)-round(cost)
        })
        
        total_box += count
        total_cost += cost
        total_profit += profit

    df_box = pd.DataFrame(user_list)
    box_path = os.path.join(folder_box, f"{date}盲盒统计.xlsx")
    df_box.to_excel(box_path, index=False)

    wb_box = load_workbook(box_path)
    ws_box = wb_box.active
    ws_box.insert_rows(1)
    ws_box.merge_cells('A1:E1')
    ws_box['A1'] = f'总盲盒数：{total_box}，共花费{round(total_cost)}电池，共开出{round(total_profit)}电池，净收益{round(total_profit)-round(total_cost)}电池'
    ws_box['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws_box.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E']:
        ws_box.column_dimensions[col].width = 15

    wb_box.save(box_path)
except FileNotFoundError:
    print("Error: no box.json")

# danmu.jsonl
folder_danmu = "analysis/danmu"
os.makedirs(folder_danmu, exist_ok=True)
total_danmu = 0

try:
    danmu_list = []
    if os.path.exists('files/danmu.jsonl'):
        with open('files/danmu.jsonl', 'r', encoding='utf-8') as f:
            for line in f:
                if line.strip():
                    danmu_list.append(json.loads(line))
                    total_danmu += 1

    if danmu_list:
        dm_stats = {} # 用于折线图
        word_dict = {} # 用于高频词

        
        # 停用词
        stop_words = ["的", "了", "在", "是", "我", "你", "他", "吧", "吗", "啊", "这", "那", "不", "就"]
        for item in danmu_list:
            # 弹幕频率
            period = (item['time'] // 60) * 60
            dm_stats[period] = dm_stats.get(period, 0) + 1
            
            # 词频
            words = jieba.lcut(item['danmu'])
            for w in words:
                if len(w) > 1 and w not in stop_words:
                    if w in word_dict:
                        word_dict[w] += 1
                    else:
                        word_dict[w] = 1

        # 弹幕频率折线图
        dm_sorted_keys = sorted(dm_stats.keys())
        dm_min_time = dm_sorted_keys[0]
        dm_max_time = dm_sorted_keys[-1]
        dm_full_times = range(dm_min_time, dm_max_time + 60, 60)
        dm_times = [datetime.fromtimestamp(t) for t in dm_full_times]
        dm_counts = [dm_stats.get(t, 0) for t in dm_full_times] 

        plt.figure(figsize=(12, 6))
        plt.plot(dm_times, dm_counts, color='#ff85c0', linewidth=2, alpha=0.8)
        plt.fill_between(dm_times, dm_counts, color='#ff85c0', alpha=0.3)
        
        ax = plt.gca()
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
        plt.title(f'{date} 云宝直播间弹幕热度密度图')
        plt.ylabel('每分钟弹幕数')
        plt.grid(axis='y', linestyle=':', alpha=0.6)
        plt.tight_layout()
        plt.savefig(f'{folder_danmu}/{date}弹幕频率.png', dpi=300)

        word_list = [{"词语": k, "总出现次数": v} for k, v in word_dict.items()]
        word_df = pd.DataFrame(word_list)
        word_df = word_df.sort_values(by="总出现次数", ascending=False)  # 降序排列

        excel_path = os.path.join(folder_danmu, f"{date}弹幕统计.xlsx")
        word_df.to_excel(excel_path, index=False)
        
        wb_dm = load_workbook(excel_path)
        ws_dm = wb_dm.active
        ws_dm.insert_rows(1)
        ws_dm.merge_cells('A1:B1')
        ws_dm['A1'] = f'直播总弹幕数：{total_danmu} 条'
        ws_dm['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_dm.column_dimensions['A'].width = 20
        ws_dm.column_dimensions['B'].width = 15
        wb_dm.save(excel_path)

except Exception as e:
    print(f"弹幕分析出错: {e}")

print("Done.")